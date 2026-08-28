from __future__ import annotations

import csv
import json
import re
import threading
import unicodedata
from dataclasses import asdict, dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from mcp_server.app.adapters.idempotency import IdempotencyStore


@dataclass(slots=True)
class ArticleSummary:
    article_id: str
    title: str
    published_at: str | None
    latest_reads: int
    last_captured_at: str | None
    status: str


class WorkspaceAdapter:
    """把现有运营仓库当作外部系统；Excel/CSV 永远只读。"""

    _write_lock = threading.Lock()

    def __init__(self, workspace: Path, approval_secret: str, state_path: Path) -> None:
        self.workspace = workspace.resolve()
        self.approval_secret = approval_secret
        self.idempotency = IdempotencyStore(state_path)

    def _require_workspace(self) -> None:
        if not self.workspace.is_dir():
            raise FileNotFoundError(f"运营项目目录不存在：{self.workspace}")

    def _read_text(self, path: Path) -> str:
        self._require_workspace()
        return path.read_text(encoding="utf-8") if path.is_file() else ""

    def playbook(self) -> str:
        return self._read_text(self.workspace / "knowledge" / "content_playbook.md")

    def recent_reviews(self, limit: int = 4) -> list[dict[str, str]]:
        items: list[tuple[float, Path]] = []
        for path in self.workspace.glob("*/weekly_review.md"):
            items.append((path.stat().st_mtime, path))
        result = []
        for _, path in sorted(items, reverse=True)[:limit]:
            result.append(
                {
                    "path": str(path.relative_to(self.workspace)),
                    "content": self._read_text(path),
                }
            )
        return result

    def recent_article_records(self, limit: int = 12) -> list[dict[str, str]]:
        items: list[tuple[float, Path]] = []
        for path in self.workspace.glob("drafts/*/article_record.md"):
            items.append((path.stat().st_mtime, path))
        result = []
        for _, path in sorted(items, reverse=True)[:limit]:
            result.append(
                {
                    "article_id": path.parent.name,
                    "path": str(path.relative_to(self.workspace)),
                    "content": self._read_text(path),
                }
            )
        return result

    @property
    def workbook_path(self) -> Path:
        return self.workspace / "data" / "wechat_realtime_metrics.xlsx"

    @staticmethod
    def _cell_datetime(value: Any) -> datetime | None:
        if isinstance(value, datetime):
            return value if value.tzinfo else value.replace(tzinfo=timezone(timedelta(hours=8)))
        if isinstance(value, str) and value.strip():
            text = value.strip().replace("Z", "+00:00")
            try:
                parsed = datetime.fromisoformat(text)
                return parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone(timedelta(hours=8)))
            except ValueError:
                return None
        return None

    def _open_workbook_read_only(self):
        self._require_workspace()
        if not self.workbook_path.is_file():
            raise FileNotFoundError(f"指标工作簿不存在：{self.workbook_path}")
        return load_workbook(self.workbook_path, read_only=True, data_only=True)

    def recent_performance(self, limit: int = 10) -> list[dict[str, Any]]:
        workbook = self._open_workbook_read_only()
        try:
            sheet = workbook["文章"]
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            rows: list[dict[str, Any]] = []
            for values in sheet.iter_rows(min_row=2, values_only=True):
                row = dict(zip(headers, values, strict=False))
                if not row.get("文章ID"):
                    continue
                published = self._cell_datetime(row.get("发布时间"))
                captured = self._cell_datetime(row.get("最后采集时间"))
                rows.append(
                    asdict(
                        ArticleSummary(
                            article_id=str(row["文章ID"]),
                            title=str(row.get("标题") or ""),
                            published_at=published.isoformat() if published else None,
                            latest_reads=int(row.get("最新阅读量") or 0),
                            last_captured_at=captured.isoformat() if captured else None,
                            status=str(row.get("状态") or ""),
                        )
                    )
                )
            rows.sort(key=lambda item: item.get("published_at") or "", reverse=True)
            return rows[:limit]
        finally:
            workbook.close()

    def hourly_curve(self, article_id: str) -> list[dict[str, Any]]:
        workbook = self._open_workbook_read_only()
        try:
            sheet = workbook["采样明细"]
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            result: list[dict[str, Any]] = []
            for values in sheet.iter_rows(min_row=2, values_only=True):
                row = dict(zip(headers, values, strict=False))
                if str(row.get("文章ID") or "") != str(article_id):
                    continue
                captured = self._cell_datetime(row.get("实际采集时间"))
                result.append(
                    {
                        "article_id": str(article_id),
                        "title": str(row.get("标题") or ""),
                        "published_at": (
                            self._cell_datetime(row.get("发布时间")).isoformat()
                            if self._cell_datetime(row.get("发布时间"))
                            else None
                        ),
                        "captured_at": captured.isoformat() if captured else None,
                        "hours_since_publish": float(row.get("发布后小时") or 0),
                        "reads": int(row.get("阅读量") or 0),
                        "shares": int(row.get("分享") or 0),
                        "likes": int(row.get("点赞") or 0),
                        "favorites": int(row.get("收藏") or 0),
                        "new_followers": int(row.get("新增关注") or 0),
                    }
                )
            result.sort(key=lambda item: item["hours_since_publish"])
            return result
        finally:
            workbook.close()

    def article_details(self, article_id: str) -> dict[str, Any]:
        detail_dir = self.workspace / "data" / "wechat_article_details"
        matches = list(detail_dir.glob(f"*{article_id}*.csv"))
        if not matches:
            # 文件名有时不含文章 ID；manifest 是第二选择。
            manifest = detail_dir / "manifest.json"
            if manifest.is_file():
                data = json.loads(manifest.read_text(encoding="utf-8"))
                for item in data if isinstance(data, list) else data.get("items", []):
                    if str(item.get("article_id")) == str(article_id) and item.get("path"):
                        candidate = detail_dir / item["path"]
                        if candidate.is_file():
                            matches = [candidate]
                            break
        if not matches:
            return {"article_id": article_id, "sections": []}
        path = max(matches, key=lambda item: item.stat().st_mtime)
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))
        return {
            "article_id": article_id,
            "source_path": str(path.relative_to(self.workspace)),
            "rows": rows,
        }

    @staticmethod
    def normalize_title(title: str) -> str:
        normalized = unicodedata.normalize("NFKC", title).lower()
        return re.sub(r"[\s\W_]+", "", normalized, flags=re.UNICODE)

    def match_articles(
        self, title: str, published_at: datetime | None, window_hours: int = 12
    ) -> list[dict[str, Any]]:
        wanted = self.normalize_title(title)
        matches = []
        for item in self.recent_performance(limit=100):
            if self.normalize_title(item["title"]) != wanted:
                continue
            if published_at and item.get("published_at"):
                actual = datetime.fromisoformat(item["published_at"])
                if abs((actual - published_at).total_seconds()) > window_hours * 3600:
                    continue
            matches.append(item)
        return matches

    def search_similar(self, query: str, limit: int = 8) -> list[dict[str, Any]]:
        query_terms = set(re.findall(r"[\w\u4e00-\u9fff]{2,}", query.lower()))
        scored = []
        for record in self.recent_article_records(limit=50):
            content = record["content"]
            terms = set(re.findall(r"[\w\u4e00-\u9fff]{2,}", content.lower()))
            score = len(query_terms & terms) / max(len(query_terms), 1)
            if score:
                scored.append({**record, "score": round(min(score, 1.0), 4)})
        return sorted(scored, key=lambda item: item["score"], reverse=True)[:limit]

    def _resolve_write_target(self, relative_path: str) -> Path:
        if Path(relative_path).is_absolute():
            raise PermissionError("写回目标必须使用运营仓库内的相对路径")
        target = (self.workspace / relative_path).resolve()
        try:
            relative = target.relative_to(self.workspace)
        except ValueError as exc:
            raise PermissionError("写回路径越过运营仓库边界") from exc
        parts = relative.parts
        allowed = (
            (len(parts) == 3 and parts[0] == "drafts" and parts[2] == "article_record.md")
            or (len(parts) == 2 and parts[1] == "weekly_review.md")
            or relative.as_posix() == "knowledge/content_playbook.md"
        )
        if not allowed or target.suffix.lower() != ".md":
            raise PermissionError(f"不允许写回该路径：{relative}")
        return target

    def approved_markdown_write(
        self,
        *,
        operation: str,
        relative_path: str,
        heading: str,
        markdown: str,
        idempotency_key: str,
        approved: bool,
        approval_secret: str,
        create_if_missing: bool = False,
    ) -> dict[str, Any]:
        if not approved or approval_secret != self.approval_secret:
            raise PermissionError("缺少有效人工批准，拒绝 Markdown 写回")
        target = self._resolve_write_target(relative_path)
        if not target.exists() and not create_if_missing:
            raise FileNotFoundError(f"写回目标不存在：{target}")
        if not self.idempotency.claim(idempotency_key, operation, str(target)):
            return {"applied": False, "duplicate": True, "path": relative_path}

        target.parent.mkdir(parents=True, exist_ok=True)
        block = f"\n\n## {heading.strip()}\n\n{markdown.strip()}\n"
        with self._write_lock:
            if not target.exists():
                target.write_text(f"# 文章记录\n{block}", encoding="utf-8")
            else:
                with target.open("a", encoding="utf-8", newline="\n") as handle:
                    handle.write(block)
        return {"applied": True, "duplicate": False, "path": relative_path}
