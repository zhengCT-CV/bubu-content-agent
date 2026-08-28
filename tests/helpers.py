from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook


def build_wechat_fixture(root: Path) -> Path:
    (root / "knowledge").mkdir(parents=True)
    (root / "5.11-5.17").mkdir(parents=True)
    (root / "drafts" / "demo-article").mkdir(parents=True)
    (root / "data" / "wechat_article_details").mkdir(parents=True)
    (root / "knowledge" / "content_playbook.md").write_text(
        "# 长期打法\n\n具体动作钩子通常比抽象说理更容易理解。", encoding="utf-8"
    )
    (root / "5.11-5.17" / "weekly_review.md").write_text(
        "# 周复盘\n\n近期身份冲突型选题的分享率更值得观察。", encoding="utf-8"
    )
    (root / "drafts" / "demo-article" / "article_record.md").write_text(
        "# 文章记录\n\n标题：别再替同事收拾烂摊子\n\n复盘：具体场景打开较好。",
        encoding="utf-8",
    )

    workbook = Workbook()
    articles = workbook.active
    articles.title = "文章"
    articles.append(
        [
            "文章ID",
            "标题",
            "发布时间",
            "后台链接",
            "状态",
            "阅读提醒阈值",
            "最新阅读量",
            "最后采集时间",
            "发现时间",
            "完成时间",
        ]
    )
    samples = workbook.create_sheet("采样明细")
    samples.append(
        [
            "采样ID",
            "文章ID",
            "标题",
            "发布时间",
            "实际采集时间",
            "发布后小时",
            "阅读量",
            "分享",
            "点赞",
            "收藏",
            "新增关注",
            "页面原始文本",
        ]
    )
    workbook.create_sheet("里程碑")
    workbook.create_sheet("提醒记录")
    workbook.create_sheet("运行日志")
    workbook.save(root / "data" / "wechat_realtime_metrics.xlsx")
    return root
