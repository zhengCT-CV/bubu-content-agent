from __future__ import annotations

import asyncio
import copy
import csv
import statistics
import threading
import time
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

LOCAL_TIMEZONE = timezone(timedelta(hours=8))


class ContentDataUnavailableError(RuntimeError):
    pass


@dataclass(slots=True)
class ContentDataSnapshot:
    signature: tuple[int, int]
    overview: dict[str, Any]
    curves: dict[str, list[dict[str, Any]]]
    articles: dict[str, dict[str, Any]]
    detail_paths: dict[str, Path]


def _datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=LOCAL_TIMEZONE)
    if isinstance(value, str) and value.strip():
        try:
            parsed = datetime.fromisoformat(value.strip().replace("Z", "+00:00"))
        except ValueError:
            return None
        return parsed if parsed.tzinfo else parsed.replace(tzinfo=LOCAL_TIMEZONE)
    return None


def _iso(value: Any) -> str | None:
    parsed = _datetime(value)
    return parsed.isoformat() if parsed else None


def _number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, str) and value.endswith("%"):
        return float(value[:-1]) / 100
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _integer(value: Any) -> int:
    return int(round(_number(value)))


def _normalize_title(value: str) -> str:
    return "".join(character.lower() for character in value if character.isalnum())


def _sheet_records(sheet) -> list[dict[str, Any]]:
    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows, ()))
    return [dict(zip(headers, values, strict=False)) for values in rows if values and values[0]]


class ContentDataService:
    """Read the automation workbook without ever mutating or locking it for writes."""

    def __init__(self, workspace: Path) -> None:
        self.workspace = workspace.resolve()
        self.workbook_path = self.workspace / "data" / "wechat_realtime_metrics.xlsx"
        self.topic_performance_path = self.workspace / "data" / "topic_performance.csv"
        self.detail_dir = self.workspace / "data" / "wechat_article_details"
        self._cache: ContentDataSnapshot | None = None
        self._lock = threading.Lock()

    async def overview(self, force: bool = False) -> dict[str, Any]:
        return await asyncio.to_thread(self._overview_sync, force)

    async def article_detail(self, article_id: str) -> dict[str, Any] | None:
        return await asyncio.to_thread(self._article_detail_sync, article_id)

    def _signature(self) -> tuple[int, int]:
        stat = self.workbook_path.stat()
        return stat.st_mtime_ns, stat.st_size

    def _overview_sync(self, force: bool) -> dict[str, Any]:
        with self._lock:
            try:
                signature = self._signature()
            except OSError as exc:
                if self._cache:
                    return self._decorate(
                        self._cache.overview, cached=True, state="updating", warning=str(exc)
                    )
                raise ContentDataUnavailableError("实时数据工作簿不存在或暂时无法访问") from exc

            if self._cache and not force and self._cache.signature == signature:
                return self._decorate(self._cache.overview, cached=True)

            last_error: Exception | None = None
            for delay in (0.0, 0.25, 0.75):
                if delay:
                    time.sleep(delay)
                try:
                    signature_before = self._signature()
                    snapshot = self._load_snapshot(signature_before)
                    if self._signature() != signature_before:
                        raise OSError("数据文件在读取过程中发生变化")
                    self._cache = snapshot
                    return self._decorate(snapshot.overview, cached=False)
                except Exception as exc:  # Excel writes can surface as several zip/xml exceptions.
                    last_error = exc

            if self._cache:
                return self._decorate(
                    self._cache.overview,
                    cached=True,
                    state="updating",
                    warning="自动化正在写入数据，当前展示上一次成功快照。",
                )
            raise ContentDataUnavailableError("实时数据正在写入或格式暂时不可读，请稍后重试") from last_error

    @staticmethod
    def _decorate(
        payload: dict[str, Any],
        *,
        cached: bool,
        state: str | None = None,
        warning: str | None = None,
    ) -> dict[str, Any]:
        result = copy.deepcopy(payload)
        result["source"]["cached"] = cached
        if state:
            result["source"]["state"] = state
        if warning:
            result["source"]["warning"] = warning
        return result

    def _load_snapshot(self, signature: tuple[int, int]) -> ContentDataSnapshot:
        workbook = load_workbook(self.workbook_path, read_only=True, data_only=True)
        try:
            articles_rows = _sheet_records(workbook["文章"])
            sample_rows = _sheet_records(workbook["采样明细"])
            milestone_rows = _sheet_records(workbook["里程碑"])
            log_rows = _sheet_records(workbook["运行日志"])
        finally:
            workbook.close()

        latest_samples: dict[str, dict[str, Any]] = {}
        curves: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in sample_rows:
            article_id = str(row.get("文章ID") or "")
            if not article_id:
                continue
            point = {
                "captured_at": _iso(row.get("实际采集时间")),
                "hours_since_publish": round(_number(row.get("发布后小时")), 3),
                "reads": _integer(row.get("阅读量")),
                "shares": _integer(row.get("分享")),
                "likes": _integer(row.get("点赞")),
                "favorites": _integer(row.get("收藏")),
                "new_followers": _integer(row.get("新增关注")),
            }
            curves[article_id].append(point)
            current = latest_samples.get(article_id)
            if not current or (
                _datetime(point["captured_at"]) or datetime.min.replace(tzinfo=LOCAL_TIMEZONE)
            ) > (_datetime(current["captured_at"]) or datetime.min.replace(tzinfo=LOCAL_TIMEZONE)):
                latest_samples[article_id] = point

        detail_paths = self._detail_index()
        articles: list[dict[str, Any]] = []
        article_index: dict[str, dict[str, Any]] = {}
        for row in articles_rows:
            article_id = str(row.get("文章ID") or "")
            if not article_id:
                continue
            latest = latest_samples.get(article_id, {})
            reads = _integer(row.get("最新阅读量"))
            published_at = _datetime(row.get("发布时间"))
            captured_at = _datetime(row.get("最后采集时间"))
            current_hours = (
                round((captured_at - published_at).total_seconds() / 3600, 3)
                if captured_at and published_at
                else _number(latest.get("hours_since_publish"))
            )
            item = {
                "article_id": article_id,
                "title": str(row.get("标题") or ""),
                "published_at": published_at.isoformat() if published_at else None,
                "url": str(row.get("后台链接") or ""),
                "status": str(row.get("状态") or ""),
                "reads": reads,
                "shares": _integer(latest.get("shares")),
                "likes": _integer(latest.get("likes")),
                "favorites": _integer(latest.get("favorites")),
                "new_followers": _integer(latest.get("new_followers")),
                "captured_at": captured_at.isoformat() if captured_at else latest.get("captured_at"),
                "hours_since_publish": current_hours,
                "share_rate": round(_number(latest.get("shares")) / reads, 6) if reads else 0,
                "like_rate": round(_number(latest.get("likes")) / reads, 6) if reads else 0,
                "has_details": _normalize_title(str(row.get("标题") or "")) in detail_paths,
            }
            articles.append(item)
            article_index[article_id] = item
            article_curve = curves.get(article_id, [])
            last_curve_point = article_curve[-1] if article_curve else None
            if (
                captured_at
                and current_hours
                > _number(last_curve_point.get("hours_since_publish") if last_curve_point else 0)
                and reads != _integer(last_curve_point.get("reads") if last_curve_point else 0)
            ):
                article_curve.append(
                    {
                        "captured_at": captured_at.isoformat(),
                        "hours_since_publish": current_hours,
                        "reads": reads,
                        "shares": None,
                        "likes": None,
                        "favorites": None,
                        "new_followers": None,
                    }
                )
                curves[article_id] = article_curve

        articles.sort(key=lambda item: item.get("published_at") or "", reverse=True)
        for points in curves.values():
            points.sort(key=lambda point: point["hours_since_publish"])

        last_captured = max(
            (_datetime(item.get("captured_at")) for item in articles if item.get("captured_at")),
            default=None,
        )
        source_state = "fresh"
        if not last_captured or datetime.now(LOCAL_TIMEZONE) - last_captured > timedelta(hours=2, minutes=15):
            source_state = "stale"

        daily: dict[str, dict[str, int]] = defaultdict(
            lambda: {"articles": 0, "reads": 0, "shares": 0, "likes": 0}
        )
        for item in articles:
            date = (item.get("published_at") or "")[:10]
            if not date:
                continue
            daily[date]["articles"] += 1
            daily[date]["reads"] += item["reads"]
            daily[date]["shares"] += item["shares"]
            daily[date]["likes"] += item["likes"]

        read_values = [item["reads"] for item in articles]
        successful_runs = sum(1 for row in log_rows if str(row.get("状态") or "").lower() == "success")
        latest_run = max(
            log_rows,
            key=lambda row: _datetime(row.get("开始时间")) or datetime.min.replace(tzinfo=LOCAL_TIMEZONE),
            default={},
        )
        historical = self._historical_baseline()
        modified_at = datetime.fromtimestamp(signature[0] / 1_000_000_000, LOCAL_TIMEZONE)
        overview = {
            "source": {
                "state": source_state,
                "cached": False,
                "warning": None,
                "file_name": self.workbook_path.name,
                "file_modified_at": modified_at.isoformat(),
                "last_captured_at": last_captured.isoformat() if last_captured else None,
                "data_version": f"{signature[0]}-{signature[1]}",
            },
            "summary": {
                "tracked_articles": len(articles),
                "total_reads": sum(read_values),
                "median_reads": round(statistics.median(read_values), 1) if read_values else 0,
                "tracking_articles": sum(1 for item in articles if item["status"] == "追踪中"),
                "completed_articles": sum(1 for item in articles if item["status"] == "已完成"),
                "sample_count": len(sample_rows),
                "milestone_count": len(milestone_rows),
                "run_count": len(log_rows),
                "collector_success_rate": round(successful_runs / len(log_rows), 4) if log_rows else 0,
                "latest_run_samples": _integer(latest_run.get("写入采样数")),
            },
            "historical_baseline": historical,
            "daily_performance": [{"date": date, **values} for date, values in sorted(daily.items())],
            "articles": articles,
        }
        return ContentDataSnapshot(signature, overview, dict(curves), article_index, detail_paths)

    def _historical_baseline(self) -> dict[str, Any] | None:
        if not self.topic_performance_path.is_file():
            return None
        try:
            with self.topic_performance_path.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))
        except OSError:
            return None
        reads = [_integer(row.get("reads")) for row in rows]
        completion_rates = [_number(row.get("completion_rate")) for row in rows if row.get("completion_rate")]
        return {
            "article_count": len(rows),
            "total_reads": sum(reads),
            "median_reads": round(statistics.median(reads), 1) if reads else 0,
            "average_completion_rate": round(statistics.mean(completion_rates), 6) if completion_rates else 0,
            "date_from": min((row.get("publish_date") or "" for row in rows), default=""),
            "date_to": max((row.get("publish_date") or "" for row in rows), default=""),
        }

    def _detail_index(self) -> dict[str, Path]:
        result: dict[str, Path] = {}
        if not self.detail_dir.is_dir():
            return result
        for path in self.detail_dir.glob("*.csv"):
            try:
                with path.open("r", encoding="utf-8-sig", newline="") as handle:
                    first_row = next(csv.reader(handle), [])
            except OSError:
                continue
            title = first_row[1].strip() if len(first_row) > 1 else ""
            if title:
                result[_normalize_title(title)] = path
        return result

    def _article_detail_sync(self, article_id: str) -> dict[str, Any] | None:
        self._overview_sync(False)
        snapshot = self._cache
        if not snapshot or article_id not in snapshot.articles:
            return None
        article = copy.deepcopy(snapshot.articles[article_id])
        detail_path = snapshot.detail_paths.get(_normalize_title(article["title"]))
        exported = self._parse_detail(detail_path) if detail_path else None
        return {
            "article": article,
            "curve": copy.deepcopy(snapshot.curves.get(article_id, [])),
            "exported_detail": exported,
        }

    @staticmethod
    def _parse_detail(path: Path) -> dict[str, Any] | None:
        try:
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.reader(handle))
        except OSError:
            return None
        if not rows:
            return None

        section_starts: dict[str, int] = {}
        for index, row in enumerate(rows):
            label = row[1].strip() if len(row) > 1 else ""
            if label in {"数据概况", "阅读转化", "阅读数据趋势明细", "性别分布", "年龄分布", "地域分布"}:
                section_starts[label] = index

        def block(name: str) -> list[list[str]]:
            start = section_starts.get(name)
            if start is None:
                return []
            result = []
            for row in rows[start + 2 :]:
                label = row[1].strip() if len(row) > 1 else ""
                if not label:
                    break
                result.append(row)
            return result

        def metrics(name: str) -> dict[str, float]:
            return {
                row[1].strip(): _number(row[2] if len(row) > 2 else 0)
                for row in block(name)
                if len(row) > 2 and row[1].strip()
            }

        trend = [
            {
                "date": row[1].strip(),
                "channel": row[2].strip(),
                "reads": _integer(row[3] if len(row) > 3 else 0),
                "shares": _integer(row[4] if len(row) > 4 else 0),
            }
            for row in block("阅读数据趋势明细")
            if len(row) > 4
        ]

        def distribution(name: str, label_key: str) -> list[dict[str, Any]]:
            return [
                {label_key: row[1].strip(), "ratio": _number(row[2] if len(row) > 2 else 0)}
                for row in block(name)
                if len(row) > 2
            ]

        channels: dict[str, int] = defaultdict(int)
        for item in trend:
            if item["channel"] != "全部":
                channels[item["channel"]] += item["reads"]
        return {
            "source_file": path.name,
            "overview": metrics("数据概况"),
            "conversion": metrics("阅读转化"),
            "daily_trend": [item for item in trend if item["channel"] == "全部"],
            "channels": [
                {"label": label, "reads": reads}
                for label, reads in sorted(channels.items(), key=lambda item: item[1], reverse=True)
            ],
            "gender": distribution("性别分布", "label"),
            "age": distribution("年龄分布", "label"),
            "regions": distribution("地域分布", "label"),
        }
